import csv
from dataclasses import dataclass
from typing import List, Optional
from openpyxl import load_workbook
import tempfile
import os


@dataclass
class TrackMetadata:
    filename_from_data: str
    track_title: str
    source_program: str
    bpm: str
    key: str
    writers: List[str]
    publishers: List[str]


def parse_spreadsheet(file_path: str) -> List[TrackMetadata]:
    """Main entry point for parsing metadata files"""
    if file_path.endswith('.xlsx'):
        return _parse_excel(file_path)
    elif file_path.endswith('.csv'):
        return _parse_csv(file_path)
    raise ValueError(f"Unsupported file format: {file_path}")


def _parse_excel(file_path: str) -> List[TrackMetadata]:
    """Parse metadata from XLSX files"""
    metadata_list = []
    print(f"\n{'=' * 40}\nParsing Excel file: {file_path}\n{'=' * 40}")

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            print(f"\nProcessing sheet: {sheet_name}")
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if len(row) < 34:
                    print(f"Skipping row {row_idx} - insufficient columns")
                    continue

                metadata = TrackMetadata(
                    filename_from_data=_safe_get(row, 0),
                    track_title=_safe_get(row, 1),
                    source_program=_safe_get(row, 2),
                    bpm=_safe_get(row, 4),
                    key=_safe_get(row, 5),
                    writers=_parse_writers(row, 12),
                    publishers=_parse_publishers(row, 18)
                )

                # Print to console
                print(f"\nTrack #{len(metadata_list) + 1}")
                print(f"File Name From Data: {metadata.filename_from_data}")
                print(f"Track Title: {metadata.track_title}")
                print(f"Source Program: {metadata.source_program}")
                print(f"BPM: {metadata.bpm} | Key: {metadata.key}")
                print(f"Writers: {', '.join(metadata.writers) or 'None'}")
                print(f"Publishers: {', '.join(metadata.publishers) or 'None'}")

                metadata_list.append(metadata)

    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}") from e

    print(f"\n{'=' * 40}")
    print(f"Total tracks parsed from Excel: {len(metadata_list)}")
    print(f"{'=' * 40}\n")
    return metadata_list


def _parse_csv(file_path: str) -> List[TrackMetadata]:
    """Parse metadata from CSV files"""
    metadata_list = []
    print(f"\n{'=' * 40}\nParsing CSV file: {file_path}\n{'=' * 40}")

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)  # Skip header row

            for row_idx, row in enumerate(reader, 1):
                if len(row) < 34:
                    print(f"Skipping row {row_idx} - insufficient columns")
                    continue

                metadata = TrackMetadata(
                    filename_from_data=_safe_get(row, 0),
                    track_title=_safe_get(row, 1),
                    source_program=_safe_get(row, 2),
                    bpm=_safe_get(row, 4),
                    key=_safe_get(row, 5),
                    writers=_parse_writers(row, 12),
                    publishers=_parse_publishers(row, 18)
                )

                # Print to console
                print(f"\nTrack #{len(metadata_list) + 1}")
                print(f"File Name From Data: {metadata.filename_from_data}")
                print(f"Track Title: {metadata.track_title}")
                print(f"Source Program: {metadata.source_program}")
                print(f"BPM: {metadata.bpm} | Key: {metadata.key}")
                print(f"Writers: {', '.join(metadata.writers) or 'None'}")
                print(f"Publishers: {', '.join(metadata.publishers) or 'None'}")

                metadata_list.append(metadata)

    except Exception as e:
        raise ValueError(f"Error parsing CSV file: {str(e)}") from e

    print(f"\n{'=' * 40}")
    print(f"Total tracks parsed from CSV: {len(metadata_list)}")
    print(f"{'=' * 40}\n")
    return metadata_list


def _parse_writers(row: list, start_idx: int) -> List[str]:
    """Parse writer information from row data"""
    writers = []
    for i in range(5):  # Max 5 writers
        offset = start_idx + (i * 10)  # 10 columns between writer groups
        if offset + 5 >= len(row):
            break

        writer = _format_writer(
            first=_safe_get(row, offset),
            middle=_safe_get(row, offset + 1),
            last=_safe_get(row, offset + 2),
            pro=_safe_get(row, offset + 3),
            cae=_safe_get(row, offset + 4),
            share=_safe_get(row, offset + 5)
        )
        if writer:
            writers.append(writer)

    return writers

def _parse_publishers(row: list, start_idx: int) -> List[str]:
    """Parse publisher information from row data"""
    publishers = []
    for i in range(5):  # Max 5 publishers
        offset = start_idx + (i * 10)  # CHANGED to 10-column spacing
        if offset + 3 >= len(row):
            break

        publisher = _format_publisher(
            name=_safe_get(row, offset),
            pro=_safe_get(row, offset + 1),
            cae=_safe_get(row, offset + 2),
            share=_safe_get(row, offset + 3)
        )
        if publisher:
            publishers.append(publisher)

    return publishers


def _format_writer(first: str, middle: str, last: str, pro: str, cae: str, share: str) -> Optional[str]:
    """Format writer information into standardized string"""
    if not any([first, middle, last]):
        return None

    components = []
    if first or middle or last:
        name = ' '.join(filter(None, [first, middle, last]))
        components.append(name.strip())
    if pro:
        components.append(f"({pro})")
    if cae:
        components.append(f"[{cae}]")
    if share:
        components.append(f"{share}%")

    return ' '.join(components) if components else None


def _format_publisher(name: str, pro: str, cae: str, share: str) -> Optional[str]:
    """Format publisher information into standardized string"""
    if not name:
        return None

    components = [name]
    if pro:
        components.append(f"({pro})")
    if cae:
        components.append(f"[{cae}]")
    if share:
        components.append(f"{share}%")

    return ' '.join(components) if components else None


def _safe_get(row: list, index: int, default: str = '') -> str:
    """Safely get value from list with fallback"""
    try:
        value = row[index]
        return str(value).strip() if value is not None else default
    except (IndexError, TypeError):
        return default